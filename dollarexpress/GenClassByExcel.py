# -*- coding: utf8 -*-

import string
import sys

import time
import xlrd
import os
from openpyxl import load_workbook

superClassName = "Sample"
factorySource = "public static SampleFactory<%s> factory = new SampleFactoryImpl<>();"

javaSourceTemplate = """
 package %s;

 import com.jjg.game.core.sample.Sample;
 import com.jjg.game.core.sample.SampleFactory;
 import com.jjg.game.core.sample.SampleFactoryImpl;
 import com.jjg.game.common.proto.ProtoDesc;
 import io.protostuff.Tag;
 import com.jjg.game.common.proto.ProtobufMessage;

/**
 * Auto generate by "Python tools"
 * @Date %s
 */
 @ProtobufMessage
 public class %s extends %s{
    %s
    public static %s get%s(int sid) {
        return (%s)factory.getSample(sid);
    }

    public static %s new%s(int sid) {
        return (%s)factory.newSample(sid);
    }
 %s
 %s
 }
"""

csSourceTemplate = """
using UnityEngine;
using System.Collections;
using System;
using System.Collections.Generic;
using ProtoBuf;

/**
 * Auto generate by "Python tools"
 * @Date %s
 */
[ProtoContract]
public class %s
{
    [ProtoMember(1, IsRequired = true)]
    public int id;
    [ProtoMember(2, IsRequired = true)]
    public String name;
 %s
}

"""

topackage = "com.jjg.game.sample"


class Field:
    comment = "field comment."
    modifier = "private"
    fieldType = "String"
    fieldName = ""
    index = 1

    def __eq__(self, o) -> bool:
        return self.fieldName == o.fieldName

    def genField(self, i):
        return "\t@Tag(%s)\n\t@ProtoDesc(\"%s\")\n\t%s %s %s;\n" % (
            self.index, self.comment, self.modifier, self.fieldType, self.fieldName)


    def genGetterSetter(self,i):
        if self.fieldType.lower() == "boolean":
            # 生成getter方法
            getter = "\tpublic %s is%s() {\n\t\treturn this.%s;\n\t}\n\n" % (
                self.fieldType,
                self.fieldName[0].upper() + self.fieldName[1:],
                self.fieldName
            )
        else:
            # 生成getter方法
            getter = "\tpublic %s get%s() {\n\t\treturn this.%s;\n\t}\n\n" % (
                self.fieldType,
                self.fieldName[0].upper() + self.fieldName[1:],
                self.fieldName
            )

        # 生成setter方法
        setter = "\tpublic void set%s(%s %s) {\n\t\tthis.%s = %s;\n\t}\n\n" % (
            self.fieldName[0].upper() + self.fieldName[1:],
            self.fieldType,
            self.fieldName,
            self.fieldName,
            self.fieldName
        )

        return getter + setter

    def genCsField(self, i):
        if self.fieldType == "String":
            self.fieldType = "string"
        elif self.fieldType == "boolean":
            self.fieldType = "bool"
        if self.fieldType == "ArrayList":
            self.fieldType = "List"

        return "\t[ProtoMember(%s, IsRequired = true)]\n\t// %s\n\t%s %s %s;\n" % (
            i + 4, self.comment, self.modifier, self.fieldType, self.fieldName)


class JavaSourceTemplate:
    package: string = ""
    spClassName = ""
    factorySrc = ""
    className = ""
    fields = []

    def __init__(self):
        self.fields = []

    def genJavaSource(self):
        strFields = ""

        setgetStr = ""
        for i, field in enumerate(self.fields):
            strFields += field.genField(i)
            setgetStr += field.genGetterSetter(i)

        return javaSourceTemplate % (
            self.package, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), self.className, self.spClassName,
            # strFields,
            self.factorySrc,
            # setgetStr,
            self.className, self.className, self.className, self.className, self.className, self.className,
            strFields,
            setgetStr
            )

    def genCsSource(self):
        strFields = ""
        for i, field in enumerate(self.fields):
            strFields += field.genCsField(i)
        return csSourceTemplate % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), self.className, strFields)

    pass


def excel2Source(parent: string, filename: string, _javaToDir: string, _csToDir: string):
    if filename.startswith("~$") or not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        return
    fullName = os.path.join(parent, filename)

    index = filename.find('.')
    pkg = filename[0: index]
    # 读取文件
    wb = load_workbook(filename=fullName, read_only=True, data_only=True)
    # package = fullName.replace(fromDir, "").split(".")[0].replace("/", ".")
#     package = topackage + "." + pkg

    for sheet in wb.worksheets:
        jst = JavaSourceTemplate()
        jst.package = topackage
        sheetName = sheet.title
        className = sheetName
        index = sheetName.find('.')
        _superClassName = superClassName
        _factorySource = factorySource % className
        if index != -1:
            _superClassName = sheetName[0: index]
            className = sheetName[index + 1:]
            _factorySource = ""

        jst.spClassName = _superClassName
        jst.factorySrc = _factorySource
        # 得到类名
        jst.className = className

        # 获取行数据
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) >= 3:
            fieldComment = rows[0]
            fieldType = rows[1]
            fieldNames = rows[2]

            for cellNum in range(len(fieldType)):
                if fieldType[cellNum] is None or fieldNames[cellNum] is None:
                    continue
                field = Field()
                field.comment = str(fieldComment[cellNum])
                field.fieldType = str(fieldType[cellNum])
                field.fieldName = str(fieldNames[cellNum])
                field.index = cellNum+1
                if field not in jst.fields and field.fieldName not in ["sid", "name"] and field.fieldType is not "":
                    jst.fields.append(field)
            pass

            if _javaToDir is not None:
                javaSource = jst.genJavaSource()
#                 print(javaSource)
                path = str(topackage.replace(".", "/")) + "/"
                sourcePath = _javaToDir + path
                if not os.path.exists(sourcePath):
                    os.makedirs(sourcePath)
                sourceFileName = sourcePath + className + ".java"
                file = open(sourceFileName, mode="w", buffering=1024, encoding="UTF-8")
                file.write(javaSource)
            if _csToDir is not None:
                csSource = jst.genCsSource()
                csSourcePath = _csToDir
                if not os.path.exists(csSourcePath):
                    os.makedirs(csSourcePath)
                cdSourceFileName = csSourcePath + className + ".cs"
                file2 = open(cdSourceFileName, mode="w", buffering=1024, encoding="UTF-8")
                file2.write(csSource)


def excelPath2Source(_fromDir: string, _javaToDir: string, _csToDir: string):
    for parent, dirnames, filenames in os.walk(_fromDir):
        for filename in filenames:
            excel2Source(parent, filename, _javaToDir, _csToDir)


if __name__ == "__main__":
    print("===================================")
    print("usages: python xxx.py fromDir javaToDir csToDir")
    print("===================================")
    fromDir = "./resources/sample/"
    toJavaDir = "./src/main/java/"
    toCsDir = None
    alen = len(sys.argv)
    if alen > 1 and sys.argv[1] is not None:
        fromDir = sys.argv[1]
    if alen > 2:
        toJavaDir = sys.argv[2]
    if alen > 3:
        toCsDir = sys.argv[3]
    print("use args,fromDir=%s,toJavaDir=%s,toCsDir=%s" % (fromDir, toJavaDir, toCsDir))
    print()
    excelPath2Source(fromDir, toJavaDir, toCsDir)
    print("生成结束")
